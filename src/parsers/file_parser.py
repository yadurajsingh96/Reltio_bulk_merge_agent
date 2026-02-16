"""
Intelligent File Parser for HCP Merge Assistant

Supports:
- CSV files
- Excel files (.xlsx, .xls)
- JSON files

Features:
- Automatic column type detection
- Intelligent mapping to HCP attributes
- Validation and cleaning
- Streaming for large files
"""

import csv
import json
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional, Iterator, Tuple
from dataclasses import dataclass, field
from enum import Enum
import re

logger = logging.getLogger(__name__)


class FileType(Enum):
    """Supported file types"""
    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"
    UNKNOWN = "unknown"


@dataclass
class ParsedRecord:
    """A single parsed record from input file"""
    row_number: int
    raw_data: Dict[str, Any]
    normalized_data: Dict[str, Any]
    identifiers: Dict[str, Any]  # NPI, entity_uri, crosswalk_id, etc.
    attributes: Dict[str, Any]   # Name, Address, Specialty, etc.
    merge_target: Optional[str] = None  # If file specifies target entity
    validation_errors: List[str] = field(default_factory=list)
    is_valid: bool = True


@dataclass
class ParsedFile:
    """Complete parsed file with metadata"""
    filename: str
    file_type: FileType
    total_records: int
    valid_records: int
    invalid_records: int
    records: List[ParsedRecord]
    column_mapping: Dict[str, str]
    detected_columns: List[str]
    warnings: List[str] = field(default_factory=list)


class ColumnMapper:
    """
    Intelligent column name mapper for HCP attributes

    Maps various column name conventions to standardized Reltio attribute names
    """

    # Standard HCP attribute mappings
    COLUMN_MAPPINGS = {
        # NPI variations
        "npi": "NPI",
        "npi_number": "NPI",
        "national_provider_identifier": "NPI",
        "provider_npi": "NPI",

        # Name variations
        "first_name": "FirstName",
        "firstname": "FirstName",
        "first": "FirstName",
        "fname": "FirstName",
        "given_name": "FirstName",

        "last_name": "LastName",
        "lastname": "LastName",
        "last": "LastName",
        "lname": "LastName",
        "family_name": "LastName",
        "surname": "LastName",

        "middle_name": "MiddleName",
        "middlename": "MiddleName",
        "middle": "MiddleName",
        "mname": "MiddleName",

        "full_name": "FullName",
        "fullname": "FullName",
        "name": "FullName",
        "provider_name": "FullName",
        "hcp_name": "FullName",

        "suffix": "Suffix",
        "name_suffix": "Suffix",

        "prefix": "Prefix",
        "title": "Prefix",
        "salutation": "Prefix",

        # Specialty variations
        "specialty": "Specialty",
        "speciality": "Specialty",
        "primary_specialty": "Specialty",
        "medical_specialty": "Specialty",
        "spec": "Specialty",

        "sub_specialty": "SubSpecialty",
        "subspecialty": "SubSpecialty",
        "secondary_specialty": "SubSpecialty",

        # Address variations
        "address": "AddressLine1",
        "address1": "AddressLine1",
        "address_line_1": "AddressLine1",
        "street": "AddressLine1",
        "street_address": "AddressLine1",

        "address2": "AddressLine2",
        "address_line_2": "AddressLine2",
        "suite": "AddressLine2",

        "city": "City",
        "town": "City",

        "state": "State",
        "state_code": "State",
        "province": "State",

        "zip": "PostalCode",
        "zipcode": "PostalCode",
        "zip_code": "PostalCode",
        "postal_code": "PostalCode",
        "postalcode": "PostalCode",

        "country": "Country",
        "country_code": "Country",

        # Contact variations
        "phone": "Phone",
        "phone_number": "Phone",
        "telephone": "Phone",
        "tel": "Phone",
        "office_phone": "Phone",

        "email": "Email",
        "email_address": "Email",
        "e_mail": "Email",

        "fax": "Fax",
        "fax_number": "Fax",

        # Identifiers
        "dea": "DEA",
        "dea_number": "DEA",

        "license": "LicenseNumber",
        "license_number": "LicenseNumber",
        "medical_license": "LicenseNumber",
        "state_license": "LicenseNumber",

        "tax_id": "TaxID",
        "taxid": "TaxID",
        "tin": "TaxID",

        # Entity references
        "entity_id": "EntityURI",
        "entity_uri": "EntityURI",
        "reltio_id": "EntityURI",
        "uri": "EntityURI",

        "source_id": "SourceID",
        "source_system_id": "SourceID",
        "crosswalk": "SourceID",
        "crosswalk_id": "SourceID",
        "external_id": "SourceID",

        # Merge targets
        "merge_with": "MergeTarget",
        "merge_target": "MergeTarget",
        "target_entity": "MergeTarget",
        "target_id": "MergeTarget",
        "merge_into": "MergeTarget",

        # Organization affiliation
        "organization": "Organization",
        "org": "Organization",
        "hospital": "Organization",
        "facility": "Organization",
        "practice": "Organization",
        "affiliation": "Organization",

        # Credentials
        "credentials": "Credentials",
        "degree": "Credentials",
        "designation": "Credentials",
    }

    # Identifier columns (used for searching in Reltio)
    IDENTIFIER_COLUMNS = {"NPI", "EntityURI", "SourceID", "DEA", "LicenseNumber", "TaxID"}

    # Columns that indicate merge target
    MERGE_TARGET_COLUMNS = {"MergeTarget"}

    @classmethod
    def normalize_column_name(cls, column: str) -> str:
        """Normalize a column name for matching"""
        # Lowercase, remove special chars, replace spaces with underscores
        normalized = column.lower().strip()
        normalized = re.sub(r'[^a-z0-9_]', '_', normalized)
        normalized = re.sub(r'_+', '_', normalized)
        normalized = normalized.strip('_')
        return normalized

    @classmethod
    def map_column(cls, column: str) -> Optional[str]:
        """Map a column name to standardized attribute name"""
        normalized = cls.normalize_column_name(column)
        return cls.COLUMN_MAPPINGS.get(normalized)

    @classmethod
    def auto_map_columns(cls, columns: List[str]) -> Dict[str, str]:
        """
        Automatically map all columns in a file

        Returns:
            Dict mapping original column name to Reltio attribute name
        """
        mapping = {}
        for col in columns:
            mapped = cls.map_column(col)
            if mapped:
                mapping[col] = mapped
            else:
                # Keep original if no mapping found
                mapping[col] = col
        return mapping

    @classmethod
    def is_identifier(cls, attr_name: str) -> bool:
        """Check if attribute is an identifier"""
        return attr_name in cls.IDENTIFIER_COLUMNS

    @classmethod
    def is_merge_target(cls, attr_name: str) -> bool:
        """Check if attribute specifies merge target"""
        return attr_name in cls.MERGE_TARGET_COLUMNS


class FileParser:
    """
    Universal file parser for CSV, Excel, and JSON

    Features:
    - Automatic file type detection
    - Intelligent column mapping
    - Data validation and cleaning
    - Support for large files via streaming
    """

    def __init__(self, column_mapping: Optional[Dict[str, str]] = None):
        """
        Initialize parser

        Args:
            column_mapping: Optional custom column mapping (original -> Reltio attribute)
        """
        self.custom_mapping = column_mapping or {}
        self.mapper = ColumnMapper()

    def detect_file_type(self, file_path: str) -> FileType:
        """Detect file type from extension"""
        path = Path(file_path)
        ext = path.suffix.lower()

        if ext == ".csv":
            return FileType.CSV
        elif ext in [".xlsx", ".xls"]:
            return FileType.EXCEL
        elif ext == ".json":
            return FileType.JSON
        else:
            return FileType.UNKNOWN

    def parse_file(
        self,
        file_path: str,
        file_type: Optional[FileType] = None,
        sheet_name: Optional[str] = None
    ) -> ParsedFile:
        """
        Parse a file and return structured records

        Args:
            file_path: Path to the file
            file_type: Optional file type (auto-detected if not provided)
            sheet_name: For Excel files, which sheet to parse

        Returns:
            ParsedFile with all records and metadata
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        # Detect file type
        if file_type is None:
            file_type = self.detect_file_type(file_path)

        if file_type == FileType.UNKNOWN:
            raise ValueError(f"Unsupported file type: {path.suffix}")

        # Parse based on type
        if file_type == FileType.CSV:
            return self._parse_csv(file_path)
        elif file_type == FileType.EXCEL:
            return self._parse_excel(file_path, sheet_name)
        elif file_type == FileType.JSON:
            return self._parse_json(file_path)

    def _parse_csv(self, file_path: str) -> ParsedFile:
        """Parse CSV file"""
        records = []
        warnings = []

        with open(file_path, 'r', encoding='utf-8-sig') as f:
            # Detect delimiter
            sample = f.read(4096)
            f.seek(0)

            try:
                dialect = csv.Sniffer().sniff(sample)
            except csv.Error:
                dialect = csv.excel

            reader = csv.DictReader(f, dialect=dialect)
            columns = reader.fieldnames or []

            # Map columns
            column_mapping = self._get_column_mapping(columns)

            for row_num, row in enumerate(reader, start=2):  # Start at 2 (header is 1)
                record = self._process_row(row_num, row, column_mapping)
                records.append(record)

        valid_count = sum(1 for r in records if r.is_valid)

        return ParsedFile(
            filename=Path(file_path).name,
            file_type=FileType.CSV,
            total_records=len(records),
            valid_records=valid_count,
            invalid_records=len(records) - valid_count,
            records=records,
            column_mapping=column_mapping,
            detected_columns=columns,
            warnings=warnings
        )

    def _parse_excel(self, file_path: str, sheet_name: Optional[str] = None) -> ParsedFile:
        """Parse Excel file"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")

        records = []
        warnings = []

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Select sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Get headers from first row
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)

        if not header_row:
            raise ValueError("Excel file appears to be empty")

        columns = [str(col) if col else f"Column_{i}" for i, col in enumerate(header_row)]
        column_mapping = self._get_column_mapping(columns)

        # Process data rows
        for row_num, row in enumerate(rows, start=2):
            row_dict = {columns[i]: row[i] for i in range(len(columns)) if i < len(row)}
            record = self._process_row(row_num, row_dict, column_mapping)
            records.append(record)

        wb.close()

        valid_count = sum(1 for r in records if r.is_valid)

        return ParsedFile(
            filename=Path(file_path).name,
            file_type=FileType.EXCEL,
            total_records=len(records),
            valid_records=valid_count,
            invalid_records=len(records) - valid_count,
            records=records,
            column_mapping=column_mapping,
            detected_columns=columns,
            warnings=warnings
        )

    def _parse_json(self, file_path: str) -> ParsedFile:
        """Parse JSON file"""
        records = []
        warnings = []

        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)

        # Handle both array and object with 'records' key
        if isinstance(data, list):
            items = data
        elif isinstance(data, dict) and 'records' in data:
            items = data['records']
        elif isinstance(data, dict):
            items = [data]  # Single record
        else:
            raise ValueError("JSON must be an array or object with 'records' key")

        # Get columns from first record
        if items:
            columns = list(items[0].keys())
            column_mapping = self._get_column_mapping(columns)
        else:
            columns = []
            column_mapping = {}

        for row_num, item in enumerate(items, start=1):
            record = self._process_row(row_num, item, column_mapping)
            records.append(record)

        valid_count = sum(1 for r in records if r.is_valid)

        return ParsedFile(
            filename=Path(file_path).name,
            file_type=FileType.JSON,
            total_records=len(records),
            valid_records=valid_count,
            invalid_records=len(records) - valid_count,
            records=records,
            column_mapping=column_mapping,
            detected_columns=columns,
            warnings=warnings
        )

    def _get_column_mapping(self, columns: List[str]) -> Dict[str, str]:
        """Get column mapping, preferring custom mapping over auto-detection"""
        mapping = ColumnMapper.auto_map_columns(columns)

        # Override with custom mappings
        for col, mapped in self.custom_mapping.items():
            if col in columns:
                mapping[col] = mapped

        return mapping

    def _process_row(
        self,
        row_num: int,
        row_data: Dict[str, Any],
        column_mapping: Dict[str, str]
    ) -> ParsedRecord:
        """Process a single row into a ParsedRecord"""
        normalized_data = {}
        identifiers = {}
        attributes = {}
        merge_target = None
        validation_errors = []

        for original_col, value in row_data.items():
            if value is None or (isinstance(value, str) and value.strip() == ""):
                continue

            # Clean value
            if isinstance(value, str):
                value = value.strip()

            # Get mapped column name
            mapped_col = column_mapping.get(original_col, original_col)
            normalized_data[mapped_col] = value

            # Categorize
            if ColumnMapper.is_identifier(mapped_col):
                identifiers[mapped_col] = value
            elif ColumnMapper.is_merge_target(mapped_col):
                merge_target = str(value)
            else:
                attributes[mapped_col] = value

        # Validation
        is_valid = True

        # Must have at least one identifier or some attributes
        if not identifiers and not attributes:
            validation_errors.append("Row has no identifiable data")
            is_valid = False

        # Validate NPI format if present
        if "NPI" in identifiers:
            npi = str(identifiers["NPI"])
            if not re.match(r'^\d{10}$', npi):
                validation_errors.append(f"Invalid NPI format: {npi}")
                # Don't mark invalid - might still be searchable

        return ParsedRecord(
            row_number=row_num,
            raw_data=row_data,
            normalized_data=normalized_data,
            identifiers=identifiers,
            attributes=attributes,
            merge_target=merge_target,
            validation_errors=validation_errors,
            is_valid=is_valid
        )

    def parse_file_streaming(
        self,
        file_path: str,
        batch_size: int = 100
    ) -> Iterator[List[ParsedRecord]]:
        """
        Parse file in streaming mode for large files

        Yields batches of records instead of loading all into memory

        Args:
            file_path: Path to file
            batch_size: Number of records per batch

        Yields:
            Batches of ParsedRecord
        """
        file_type = self.detect_file_type(file_path)

        if file_type == FileType.CSV:
            yield from self._stream_csv(file_path, batch_size)
        elif file_type == FileType.EXCEL:
            yield from self._stream_excel(file_path, batch_size)
        elif file_type == FileType.JSON:
            # JSON requires full load, but we can batch the output
            parsed = self._parse_json(file_path)
            for i in range(0, len(parsed.records), batch_size):
                yield parsed.records[i:i + batch_size]

    def _stream_csv(
        self,
        file_path: str,
        batch_size: int
    ) -> Iterator[List[ParsedRecord]]:
        """Stream CSV file in batches"""
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            columns = reader.fieldnames or []
            column_mapping = self._get_column_mapping(columns)

            batch = []
            for row_num, row in enumerate(reader, start=2):
                record = self._process_row(row_num, row, column_mapping)
                batch.append(record)

                if len(batch) >= batch_size:
                    yield batch
                    batch = []

            if batch:
                yield batch

    def _stream_excel(
        self,
        file_path: str,
        batch_size: int
    ) -> Iterator[List[ParsedRecord]]:
        """Stream Excel file in batches"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl required for Excel files")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)

        if not header_row:
            wb.close()
            return

        columns = [str(col) if col else f"Column_{i}" for i, col in enumerate(header_row)]
        column_mapping = self._get_column_mapping(columns)

        batch = []
        for row_num, row in enumerate(rows, start=2):
            row_dict = {columns[i]: row[i] for i in range(len(columns)) if i < len(row)}
            record = self._process_row(row_num, row_dict, column_mapping)
            batch.append(record)

            if len(batch) >= batch_size:
                yield batch
                batch = []

        if batch:
            yield batch

        wb.close()
